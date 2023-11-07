from django.shortcuts import render,redirect
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.urls import reverse
from django.contrib import messages
from .models import Comment, Project, Ticket, Type, Priority, Status
from accounts.models import EmployeeProfile, Role
from .forms import CommentForm, EditTicketForm, CreateProjectFormAdmin, EditProjectFormAdmin, CreateTicketForm,  ManageRoleAssignment, EditManageRoleAssignment, EditTicketFormDev
from django.contrib.auth.models import User
from django.db.models import Q
from django.db.models import Count
from .tables import ProjectTable
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import *
from .decorators import role_required
import csv
import datetime
from datetime import date

def is_valid_queryparam(param):
  return param != '' and param is not None


@login_required(login_url="/")
def home(request):
  labels = []
  projects_data = []
  tickets_type = []
  num_tickets = []
  projects = Project.objects.all()
  projects_table = ProjectTable()
  projects_tickets = Ticket.objects.filter()
  tickets = Ticket.objects.all()
  tickets_with_project_counts = Project.objects.annotate(ticket_count=Count('ticket'))

  tickets_type_issue_count = Ticket.objects.filter(type__type="issue").count()
  tickets_type_bug_count = Ticket.objects.filter(type__type="bug").count()
  tickets_type_featurerequired_count = Ticket.objects.filter(type__type="feature required").count()

  tickets_priority_low_count = Ticket.objects.filter(priority__priority="low").count()
  tickets_priority_intermediate_count = Ticket.objects.filter(priority__priority="intermediate").count()
  tickets_priority_high_count = Ticket.objects.filter(priority__priority="high").count()

  tickets_priority_open_count = Ticket.objects.filter(status__status="open").count()
  tickets_priority_inprogress_count = Ticket.objects.filter(status__status="in progress").count()
  tickets_priority_resolved_count = Ticket.objects.filter(status__status="resolved").count()
  tickets_priority_closed_count = Ticket.objects.filter(status__status="closed").count()

  tickets_priorities_counts = Ticket.objects.annotate(ticket_priorities_count=Count('priority'))
  tickets_status_counts = Ticket.objects.annotate(ticket_status_count=Count('status'))
  types = Type.objects.all()
  priorities = Priority.objects.all()
  status = Status.objects.all()
  for project in projects:
    labels.append(project.name)
    projects_data.append(project.name)
    num_tickets.append(project.name)
  context={
    'projects': projects,
    'projects_data': projects_data,
    'labels': labels,
    'projects_data': projects_data,
    'projects_table': projects_table,
    'tickets':tickets,
    'types':types,
    'priorities':priorities,
    'status':status,
    'tickets_with_project_counts':tickets_with_project_counts,
    'tickets_type_issue_count':tickets_type_issue_count,
    'tickets_type_bug_count':tickets_type_bug_count,
    'tickets_type_featurerequired_count':tickets_type_featurerequired_count,
    'tickets_priority_low_count':tickets_priority_low_count,
    'tickets_priority_intermediate_count':tickets_priority_intermediate_count,
    'tickets_priority_high_count':tickets_priority_high_count,
    'tickets_priority_open_count':tickets_priority_open_count,
    'tickets_priority_inprogress_count':tickets_priority_inprogress_count,
    'tickets_priority_resolved_count':tickets_priority_resolved_count,
    'tickets_priority_closed_count':tickets_priority_closed_count,

    
  }
  return render(request, 'bugtrackerapp/home.html', context)


@login_required(login_url="/")
@role_required(["Admin"])
def manage_role_assignment(request):
  employees = EmployeeProfile.objects.all()
  manage_role_form = ManageRoleAssignment(request.POST or None, request.FILES or None)
  if request.method == 'POST':
    manage_role_form = ManageRoleAssignment(request.POST or None, request.FILES or None)
    if manage_role_form.is_valid():
      manage_role_form.save()
      # first_employee = EmployeeProfile.objects.order_by('-created')[0]
      # print(first_employee)
      messages.success(request, 'Succesfully assigned role')
      #messages.success(request, f'Succesfully assigned a role to {roleassignment.user} as a {roleassignment.role}')
      return redirect('bugtracker:manage_role_assignment')
  context = {
    'manage_role_form':manage_role_form,
    'employees': employees
  }
  return render(request, 'bugtrackerapp/add_user_roles.html', context)

@role_required(allowed_roles=["Admin"])
@login_required(login_url="/")
def role_update(request, roleassignment_id):
  roleassignment = get_object_or_404(EmployeeProfile, id=roleassignment_id)
  roleupdate_form = EditManageRoleAssignment(instance=roleassignment)
  if request.method == 'POST':
    roleupdate_form = EditManageRoleAssignment(request.POST or None, request.FILES or None, instance=roleassignment)
    if roleupdate_form.is_valid():
      roleupdate_form.save()
      messages.success(request, f'Succesfully assigned a role to {roleassignment.user} as a {roleassignment.role}')
      return redirect('bugtracker:manage_role_assignment')
  return render(request, 'bugtrackerapp/role_update.html', {'roleupdate_form': roleupdate_form, 'roleassignment':roleassignment})

@login_required(login_url="/")
def projects_view(request):
  projects = Project.objects.all().order_by('-created')
  return render(request, 'bugtrackerapp/projects.html', {'projects':projects})

@login_required(login_url="/")
def project_detail(request, project_id):
  project = get_object_or_404(Project, id=project_id)
  tickets_by_project = Ticket.objects.filter(project=project)
  tickets_by_project_count = tickets_by_project.count()
  context = {
    'project':project,
    'tickets_by_project':tickets_by_project,
    'tickets_by_project_count':tickets_by_project_count,
  }
  return render(request, 'bugtrackerapp/project_detail.html', context)

@role_required(allowed_roles=["Admin"])
@login_required(login_url="/")
def add_project(request):
  add_project_form = CreateProjectFormAdmin(request.POST or None, request.FILES or None)
  if request.method == 'POST':
    add_project_form = CreateProjectFormAdmin(request.POST or None, request.FILES or None)
    if add_project_form.is_valid():
      add_project_form.save()
      return redirect('bugtracker:projects_view')
  return render(request, 'bugtrackerapp/add_project.html', {'add_project_form':add_project_form})

@login_required(login_url="/")
@role_required(allowed_roles=["Admin", "Tester"])
def add_ticket(request):
  add_ticket_form = CreateTicketForm(request.POST or None, request.FILES or None)
  if request.method == 'POST':
    add_ticket_form = CreateTicketForm(request.POST or None, request.FILES or None)
    if add_ticket_form.is_valid():
      add_ticket_form.save()
      return redirect('bugtracker:projects_view')
  return render(request, 'bugtrackerapp/add_ticket.html', {'add_ticket_form':add_ticket_form})

@login_required(login_url="/")
def tickets_view(request):
  #display all tickets except resolved status
  #tickets = Ticket.objects.exclude(status=3).order_by('-created')
  tickets = Ticket.objects.all().order_by('-created')
  resolved_tickets = Ticket.objects.filter(status=3)
  tickets_count = tickets.count()
  resolved_tickets_count = resolved_tickets.count()
  history = Ticket.history.all()
  context = {
    'resolved_tickets':resolved_tickets,
    'tickets':tickets,
    'tickets_count':tickets_count,
    'resolved_tickets_count':resolved_tickets_count,
    'history': history
  }
  return render(request, 'bugtrackerapp/tickets.html', context)

def tickets_export_to_csv(request):
  today = date.today()
  filename_csv = f'ticket {today}.csv'
  # we wanna return text file
  response = HttpResponse(content_type='text/csv')
  response['Content-Disposition'] = f'attachment; filename={filename_csv}'
  # Create a csv writer
  writer = csv.writer(response)
  # Designate the Model
  tickets = Ticket.objects.order_by('title')
  # Add column headings to the csv file
  writer.writerow(['Title', 'Description', 'Status', 'Author', 'Created Date'])
  # Loop tru and output
  for ticket in tickets:
    writer.writerow([ticket.title, ticket.ticket_description, ticket.status, ticket.author, ticket.created])
  return response


def tickets_export_to_xlsx(request):
  ticket_queryset = Ticket.objects.all()
  response = HttpResponse(
      content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
  )
  response['Content-Disposition'] = 'attachment; filename=tickets.xlsx'
  workbook = Workbook()
  # Get active worksheet/tab
  worksheet = workbook.active
  worksheet.title = 'Movies'
  # Define the titles for columns
  columns = [
      'Title',
      'Description',
      'Status',
      'Author',
      'Created Date'
  ]
  row_num = 1
  # Assign the titles for each cell of the header
  for col_num, column_title in enumerate(columns, 1):
      cell = worksheet.cell(row=row_num, column=col_num)
      cell.value = column_title
  # Iterate through all movies
  for ticket in ticket_queryset:
      row_num += 1
      # Define the data for each cell in the row 
      row = [
        ticket.title,
        ticket.ticket_description,
        str(ticket.status),
        str(ticket.author),
        ticket.created.strftime('%d-%m-%Y')
      ]
      # Assign the data for each cell of the row 
      for col_num, cell_value in enumerate(row, 1):
          cell = worksheet.cell(row=row_num, column=col_num)
          cell.value = cell_value

  workbook.save(response)

  return response

@login_required(login_url="/")
def ticket_detail(request, ticket_id):
  comment = ''
  comment_user = ''
  comment_form = CommentForm(request.POST or None)
  ticket = get_object_or_404(Ticket, id=ticket_id)
  edit_ticket_dev_form = EditTicketFormDev(instance=ticket)
  ticket_developer = Ticket.objects.filter(assign_members=request.user)
  # show/hide buttons based on current user
  is_user_ticket = ticket.author == request.user.profile
  is_developer_ticket = ticket.assign_members.contains(request.user)
  is_tester_ticket = ticket.assign_members.contains(request.user)
  # comment
  comments = Comment.objects.filter(ticket=ticket).order_by('-date')
  # Ticket update form
  if request.method == "POST":
    edit_ticket_dev_form = EditTicketFormDev(request.POST or None, instance=ticket)
    comment_form = CommentForm(request.POST or None)
    if edit_ticket_dev_form.is_valid():
      edit_ticket_dev_form.save()
      messages.success(request, 'Status Updated')
      return redirect(reverse('bugtracker:ticket_detail', args=[ticket.id]))
  #create comment form
    if comment_form.is_valid():
      comment = comment_form.save(commit=False)
      comment.ticket = ticket
      comment.name = request.user.profile
      comment.save()
      for comment in comments:
        comment_user = comment.name
      print(comment_user)
      return redirect(reverse('bugtracker:ticket_detail', args=[ticket.id]))
  context = {
  'ticket':ticket, 
  'is_user_ticket':is_user_ticket,
  'is_developer_ticket':is_developer_ticket,
  'is_tester_ticket':is_tester_ticket,
  'comments': comments,
  'comment_user': comment_user,
  'comment_form':comment_form,
  'edit_ticket_dev_form':edit_ticket_dev_form
  }
  return render(request, 'bugtrackerapp/ticket_detail.html', context)

def comment_delete(request, comment_id):
  #ticket = get_object_or_404(Ticket, id=ticket_id)
  comment = get_object_or_404(Comment, id=comment_id)
  #ticket_project_id = ticket.project.id
  comment.delete()
  #return render(request, 'bugtrackerapp/ticket_detail.html')
  return redirect(reverse('bugtracker:ticket_detail', args=[comment.ticket.id]))

def history_delete(request):
  ticket = Ticket.history.all()
  #history_records = ticket.history.all()
  #history_record = history_records[0] 
  ticket.delete()
  return redirect('bugtracker:tickets_view')

@login_required(login_url="/")
def ticket_update(request, ticket_id):
  ticket = get_object_or_404(Ticket, id=ticket_id)
  ticket_form = EditTicketForm(instance=ticket)
  ticket_form_dev = EditTicketForm(instance=ticket)
  ticket_project_id = ticket.project.id
  if request.method == 'POST':
    ticket_form = EditTicketForm(request.POST or None, request.FILES or None, instance=ticket)
    if ticket_form.is_valid():
      ticket_form.save()
      #return redirect(reverse('bugtracker:project_detail', args=[ticket_project_id]))
      return redirect('bugtracker:tickets_view')
  # only admin can edit project field
  if not request.user.profile.role.role == "Admin":
    #ticket_form.fields["project"].disabled = True
    ticket_form.fields["project"].widget.attrs['readonly'] = True
  # only admin and tester can edit assign a ticket
  # if not request.user.profile.role.role == "Admin" and not request.user.profile.role.role == "Tester":
    # ticket_form.fields["title"].disabled = True
    # ticket_form.fields["ticket_description"].disabled = True
    # ticket_form.fields["image"].disabled = True
    #ticket_form.fields["assign_members"].disabled = True
    # ticket_form.fields["type"].disabled = True
    # ticket_form.fields["priority"].disabled = True
  return render(request, 'bugtrackerapp/ticket_update.html', {'ticket_form': ticket_form})

@login_required(login_url="/")
def ticket_delete(request, ticket_id):
  ticket = get_object_or_404(Ticket, id=ticket_id)
  ticket_project_id = ticket.project.id
  ticket.delete()
  messages.warning(request, 'Ticket Deleted')
  return redirect(reverse('bugtracker:project_detail', args=[ticket_project_id]))

@login_required(login_url="/")
def tickets_tester(request):
    employee = EmployeeProfile.objects.get(user=request.user)
    tickets = Ticket.objects.filter(
      Q(assign_members=request.user) | Q(author=employee)
    )
    context = {
      'tickets':tickets
    }
    print(tickets)
    return render(request, 'bugtrackerapp/tickets_tester.html', context)

@login_required(login_url="/")
def tickets_developer(request):
    employee = EmployeeProfile.objects.get(user=request.user)
    tickets = Ticket.objects.filter(
      Q(assign_members=request.user) | Q(author=employee)
    )
    context = {
      'tickets':tickets
    }
    return render(request, 'bugtrackerapp/tickets_developer.html', context)

def reports(request):
    return render(request, 'bugtrackerapp/reports.html', {})

def about(request):
    return render(request, 'bugtrackerapp/about.html', {})